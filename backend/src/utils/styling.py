"""
Excel styling constants — single source of truth for all openpyxl
PatternFill, Font, Border, and Alignment objects used in report generation.

Import from here instead of defining fills/fonts inline in other modules.
"""

from openpyxl.styles import Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Fills
# ---------------------------------------------------------------------------

DATA_HEADER_FILL = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
UNITS_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SEPARATOR_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
PARAM_HEADER_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
PARAM_UNIT_FILL = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")

# ---------------------------------------------------------------------------
# Fonts
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, size=11)
DATA_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
UNITS_FONT = Font(italic=True, size=10)
TITLE_FONT = Font(bold=True, size=14)

# ---------------------------------------------------------------------------
# Borders
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
