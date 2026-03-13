"""
XLSX Report Builder - Creates 4-sheet workbooks per the redesigned spec.

Sheets:
1. Data - selected original columns + derived columns
2. Calculations - derived column definitions
3. Analysis - metrics and parameters
4. Charts - embedded PNG images from matplotlib
"""

import logging
from dataclasses import dataclass
from io import BytesIO
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.utils.helpers import normalize_hex_color as _normalize_fill_color
from src.utils.styling import (
    DATA_HEADER_FILL,
    DATA_HEADER_FONT,
    HEADER_FILL,
    HEADER_FONT,
    PARAM_HEADER_FILL,
    PARAM_UNIT_FILL,
    SEPARATOR_FILL,
    THIN_BORDER,
    TITLE_FONT,
    UNITS_FILL,
    UNITS_FONT,
)

logger = logging.getLogger(__name__)


@dataclass
class DerivedColumnSpec:
    """Spec for a derived column in the report."""

    name: str
    formula: str
    unit: str = ""
    description: str = ""
    dependencies: str = ""
    enabled: bool = True


@dataclass
class AnalysisMetricSpec:
    """Spec for an analysis metric in the report."""

    chart_id: str
    chart_title: str
    name: str
    value: float
    unit: str = ""
    x_column: str = ""
    y_column: str = ""


@dataclass
class ParameterSpec:
    """Spec for a parameter in the report."""

    name: str
    value: float
    unit: str = ""


class XlsxReportBuilder:
    """
    Builds a 4-sheet XLSX report per the redesigned specification.
    """

    def build(
        self,
        df: pd.DataFrame,
        column_units: dict[str, str],
        derived_columns: list[DerivedColumnSpec],
        metrics: list[AnalysisMetricSpec],
        parameters: list[ParameterSpec],
        chart_images: list[tuple],  # List of (title, BytesIO_png)
        project_name: str = "Report",
        column_layout: dict[str, Any] | None = None,
        column_colors: dict[str, str] | None = None,
        header_mapping: dict[str, str] | None = None,
    ) -> BytesIO:
        """
        Build the complete 4-sheet XLSX report.

        Args:
            df: DataFrame with all data (selected originals + derived)
            column_units: Mapping of column name -> unit string
            derived_columns: Derived column definitions
            metrics: Analysis metrics
            parameters: Parameters used
            chart_images: List of (title, BytesIO) for chart PNGs
            project_name: Name for the report
            column_layout: Optional dict with column_order, linked_groups, separator_indices
            column_colors: Optional dict of column name -> hex color string
            header_mapping: Optional dict of column key -> display label

        Returns:
            BytesIO containing the XLSX file
        """
        logger.info(f"Building XLSX report: {project_name}")

        # Apply column reordering if specified
        if column_layout and column_layout.get("column_order"):
            ordered_cols = [c for c in column_layout["column_order"] if c in df.columns]
            # Add any columns not in the order (derived columns added later, etc.)
            for c in df.columns:
                if c not in ordered_cols:
                    ordered_cols.append(c)
            df = df[ordered_cols]

        wb = Workbook()

        # Sheet 1: Data (+ area results & chart thumbnails below data)
        separator_indices = column_layout.get("separator_indices", []) if column_layout else []
        separator_color = column_layout.get("separator_color") if column_layout else None
        self._build_data_sheet(
            wb,
            df,
            column_units,
            parameters,
            separator_indices,
            separator_color,
            column_colors or {},
            header_mapping or {},
            metrics=metrics,
            chart_images=chart_images,
            project_name=project_name,
        )

        # Sheet 2: Calculations (+ parameter/area sidebar)
        self._build_calculations_sheet(wb, derived_columns, parameters, metrics)

        # Sheet 3: Analysis (print-friendly redesign)
        self._build_analysis_sheet(wb, metrics, parameters, project_name)

        # Sheet 4: Charts (+ parameter/area context header)
        self._build_charts_sheet(wb, chart_images, parameters, metrics)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info("XLSX report built successfully")
        return output

    def _build_data_sheet(
        self,
        wb: Workbook,
        df: pd.DataFrame,
        column_units: dict[str, str],
        parameters: list[ParameterSpec],
        separator_indices: list[int],
        separator_color: str | None,
        column_colors: dict[str, str],
        header_mapping: dict[str, str],
        metrics: list[AnalysisMetricSpec] | None = None,
        chart_images: list[tuple] | None = None,
        project_name: str = "Report",
    ) -> None:
        """Sheet 1: Data - title, headers, units, data rows, separators, parameter metadata,
        area results, and chart thumbnails."""
        ws = wb.active
        ws.title = "Data"
        normalized_separator = _normalize_fill_color(separator_color)
        separator_fill = (
            PatternFill(
                start_color=normalized_separator, end_color=normalized_separator, fill_type="solid"
            )
            if normalized_separator
            else SEPARATOR_FILL
        )

        columns = list(df.columns)
        column_group_fills: dict[str, PatternFill] = {}
        for col_name in columns:
            normalized = _normalize_fill_color(column_colors.get(col_name))
            if normalized:
                column_group_fills[col_name] = PatternFill(
                    start_color=normalized,
                    end_color=normalized,
                    fill_type="solid",
                )

        # Build column list with separator columns inserted
        write_columns: list[str | None] = []  # None = separator column
        sorted_seps = sorted(set(separator_indices))
        sep_set = set(sorted_seps)

        for i, col_name in enumerate(columns):
            write_columns.append(col_name)
            if i in sep_set:
                write_columns.append(None)  # separator

        num_write_cols = len(write_columns)

        # Sheet title in row 1 spanning data + parameter sidebar
        last_col_for_title = num_write_cols + (2 + len(parameters) if parameters else 0)
        last_col_for_title = max(1, last_col_for_title)
        title_cell = ws.cell(row=1, column=1, value=project_name)
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=last_col_for_title,
        )
        ws.row_dimensions[1].height = 28

        # Row 2: Headers
        for col_idx, col_name in enumerate(write_columns, 1):
            if col_name is None:
                # Separator column
                cell = ws.cell(row=2, column=col_idx, value="")
                cell.fill = separator_fill
                cell.border = THIN_BORDER
            else:
                display_name = header_mapping.get(col_name, col_name)
                cell = ws.cell(row=2, column=col_idx, value=display_name)
                cell.font = DATA_HEADER_FONT
                cell.fill = DATA_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

        # Row 3: Units
        for col_idx, col_name in enumerate(write_columns, 1):
            if col_name is None:
                cell = ws.cell(row=3, column=col_idx, value="")
                cell.fill = separator_fill
                cell.border = THIN_BORDER
            else:
                unit = column_units.get(col_name, "")

                cell = ws.cell(row=3, column=col_idx, value=unit)
                cell.font = UNITS_FONT
                cell.fill = column_group_fills.get(col_name, UNITS_FILL)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

        # Row 4+: Data
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, col_name in enumerate(write_columns, 1):
                if col_name is None:
                    # Separator column - green fill, empty
                    cell = ws.cell(row=row_idx, column=col_idx, value="")
                    cell.fill = separator_fill
                else:
                    value = row[col_name]
                    if pd.isna(value):
                        cell = ws.cell(row=row_idx, column=col_idx, value="")
                    elif isinstance(value, int | np.integer):
                        cell = ws.cell(row=row_idx, column=col_idx, value=int(value))
                        cell.number_format = "0"
                    elif isinstance(value, float | np.floating):
                        float_val = float(value)
                        if not np.isfinite(float_val):
                            cell = ws.cell(row=row_idx, column=col_idx, value="")
                        else:
                            cell = ws.cell(row=row_idx, column=col_idx, value=float_val)
                            cell.number_format = "0.000000"
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx, value=str(value))

                    group_fill = column_group_fills.get(col_name)
                    if group_fill:
                        cell.fill = group_fill

        # Auto-width columns
        for col_idx, col_name in enumerate(write_columns, 1):
            col_letter = get_column_letter(col_idx)
            if col_name is None:
                ws.column_dimensions[col_letter].width = 3  # Narrow separator
            else:
                ws.column_dimensions[col_letter].width = max(12, len(col_name) + 4)

        # Parameter metadata area: 2 columns gap after data, then vertical block
        if parameters:
            param_start_col = num_write_cols + 3  # 2-column gap

            # Row 2: Parameter names (bold, gray bg)
            for p_idx, p in enumerate(parameters):
                col = param_start_col + p_idx
                cell = ws.cell(row=2, column=col, value=p.name)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

            # Row 3: Units (italic, light gray)
            for p_idx, p in enumerate(parameters):
                col = param_start_col + p_idx
                cell = ws.cell(row=3, column=col, value=p.unit)
                cell.font = UNITS_FONT
                cell.fill = PARAM_UNIT_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

            # Row 4: Values
            for p_idx, p in enumerate(parameters):
                col = param_start_col + p_idx
                if isinstance(p.value, float | np.floating) and not np.isfinite(float(p.value)):
                    cell = ws.cell(row=4, column=col, value="")
                else:
                    cell = ws.cell(row=4, column=col, value=p.value)
                    if isinstance(p.value, float):
                        cell.number_format = "0.000000"
                cell.border = THIN_BORDER

            # Set widths for parameter columns
            for p_idx, p in enumerate(parameters):
                col_letter = get_column_letter(param_start_col + p_idx)
                ws.column_dimensions[col_letter].width = max(12, len(p.name) + 4)

        # ── Area results in the sidebar (right side, below parameters, always visible) ──
        area_metrics = [m for m in (metrics or []) if self._is_area_metric(m)]
        sidebar_col = num_write_cols + 3  # same starting column as parameters
        sidebar_row = 6  # below param rows (2=names, 3=units, 4=values, 5=gap)

        if area_metrics:
            cell = ws.cell(row=sidebar_row, column=sidebar_col, value="Area Results")
            cell.font = TITLE_FONT
            end_col = min(sidebar_col + 2, sidebar_col + max(len(parameters) - 1, 2))
            ws.merge_cells(
                start_row=sidebar_row, start_column=sidebar_col,
                end_row=sidebar_row, end_column=end_col,
            )
            sidebar_row += 1

            for col_offset, header in enumerate(["Metric", "Value", "Unit"]):
                cell = ws.cell(row=sidebar_row, column=sidebar_col + col_offset, value=header)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            sidebar_row += 1

            for m in area_metrics:
                name_cell = ws.cell(row=sidebar_row, column=sidebar_col, value=m.name)
                name_cell.font = Font(bold=True, size=11)
                name_cell.border = THIN_BORDER
                if isinstance(m.value, float | np.floating) and not np.isfinite(float(m.value)):
                    ws.cell(row=sidebar_row, column=sidebar_col + 1, value="").border = THIN_BORDER
                else:
                    val_cell = ws.cell(row=sidebar_row, column=sidebar_col + 1, value=m.value)
                    if isinstance(m.value, float):
                        val_cell.number_format = "0.000000"
                    val_cell.border = THIN_BORDER
                ws.cell(row=sidebar_row, column=sidebar_col + 2, value=m.unit).border = THIN_BORDER
                sidebar_row += 1

            sidebar_row += 1  # spacer before charts

        # Ensure sidebar column widths cover at least 3 columns for area table
        for offset, w in enumerate([18, 16, 12]):
            col_letter = get_column_letter(sidebar_col + offset)
            existing = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(existing, w)

        # ── Chart thumbnails in the sidebar (right side, below area results) ──
        images = chart_images or []
        if images:
            THUMB_WIDTH = 520
            THUMB_HEIGHT = 325
            ROWS_PER_THUMB = 22

            for chart_title, chart_buf in images:
                try:
                    chart_buf.seek(0)
                    buf_copy = BytesIO(chart_buf.read())
                    chart_buf.seek(0)
                    img = XlImage(buf_copy)
                    img.width = THUMB_WIDTH
                    img.height = THUMB_HEIGHT
                    anchor_col = get_column_letter(sidebar_col)
                    ws.add_image(img, f"{anchor_col}{sidebar_row}")
                    sidebar_row += ROWS_PER_THUMB
                except Exception as e:
                    logger.warning(f"Data sheet chart thumbnail '{chart_title}' failed: {e}")
                    ws.cell(row=sidebar_row, column=sidebar_col, value=f"Chart unavailable: {e}")
                    sidebar_row += 2

        logger.debug(
            f"Data sheet: {len(df)} rows x {len(columns)} columns "
            f"({len(separator_indices)} separators, {len(parameters)} params, "
            f"{len(area_metrics)} area metrics, {len(images)} charts)"
        )

    @staticmethod
    def _is_area_metric(metric: AnalysisMetricSpec) -> bool:
        """Heuristic check for area-related metrics."""
        name_lower = metric.name.lower()
        return "area" in name_lower or "integral" in name_lower or "energy" in name_lower

    def _write_param_area_sidebar(
        self,
        ws,
        parameters: list[ParameterSpec],
        metrics: list[AnalysisMetricSpec],
        start_col: int,
        start_row: int = 1,
    ) -> None:
        """Write a compact parameters + area-results sidebar starting at the given column."""
        if not parameters and not metrics:
            return

        area_metrics = [m for m in metrics if self._is_area_metric(m)]
        row = start_row

        # Parameters block
        if parameters:
            cell = ws.cell(row=row, column=start_col, value="Parameters")
            cell.font = TITLE_FONT
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=start_col + 2,
            )
            row += 1

            for col_offset, header in enumerate(["Name", "Value", "Unit"]):
                cell = ws.cell(row=row, column=start_col + col_offset, value=header)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for p in parameters:
                ws.cell(row=row, column=start_col, value=p.name).border = THIN_BORDER
                if isinstance(p.value, float | np.floating) and not np.isfinite(float(p.value)):
                    ws.cell(row=row, column=start_col + 1, value="").border = THIN_BORDER
                else:
                    cell = ws.cell(row=row, column=start_col + 1, value=p.value)
                    if isinstance(p.value, float):
                        cell.number_format = "0.000000"
                    cell.border = THIN_BORDER
                ws.cell(row=row, column=start_col + 2, value=p.unit).border = THIN_BORDER
                row += 1

        # Area results block
        if area_metrics:
            row += 1
            cell = ws.cell(row=row, column=start_col, value="Area Results")
            cell.font = TITLE_FONT
            ws.merge_cells(
                start_row=row, start_column=start_col,
                end_row=row, end_column=start_col + 2,
            )
            row += 1

            for col_offset, header in enumerate(["Metric", "Value", "Unit"]):
                cell = ws.cell(row=row, column=start_col + col_offset, value=header)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for m in area_metrics:
                cell = ws.cell(row=row, column=start_col, value=m.name)
                cell.font = Font(bold=True, size=11)
                cell.border = THIN_BORDER
                if isinstance(m.value, float | np.floating) and not np.isfinite(float(m.value)):
                    ws.cell(row=row, column=start_col + 1, value="").border = THIN_BORDER
                else:
                    val_cell = ws.cell(row=row, column=start_col + 1, value=m.value)
                    if isinstance(m.value, float):
                        val_cell.number_format = "0.000000"
                    val_cell.border = THIN_BORDER
                ws.cell(row=row, column=start_col + 2, value=m.unit).border = THIN_BORDER
                row += 1

        # Column widths for sidebar
        for offset, w in enumerate([18, 16, 12]):
            ws.column_dimensions[get_column_letter(start_col + offset)].width = w

    def _build_calculations_sheet(
        self,
        wb: Workbook,
        derived_columns: list[DerivedColumnSpec],
        parameters: list[ParameterSpec] | None = None,
        metrics: list[AnalysisMetricSpec] | None = None,
    ) -> None:
        """Sheet 2: Calculations - print-friendly layout matching Analysis tab style."""
        ws = wb.create_sheet("Calculations")
        calc_col_count = 6
        active_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        disabled_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

        # ── Title row ──
        row = 1
        title_cell = ws.cell(row=row, column=1, value="Derived Columns & Calculations")
        title_cell.font = Font(bold=True, size=16, color="2F5597")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=calc_col_count,
        )
        ws.row_dimensions[row].height = 32
        row += 2  # blank spacer

        # ── Parameters section (compact, at the top like Analysis tab) ──
        params = parameters or []
        if params:
            section_cell = ws.cell(row=row, column=1, value="Parameters")
            section_cell.font = Font(bold=True, size=13, color="2F5597")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.row_dimensions[row].height = 24
            row += 1

            for col_offset, header in enumerate(["Name", "Value", "Unit"]):
                cell = ws.cell(row=row, column=1 + col_offset, value=header)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for p in params:
                ws.cell(row=row, column=1, value=p.name).border = THIN_BORDER
                if isinstance(p.value, float | np.floating) and not np.isfinite(float(p.value)):
                    ws.cell(row=row, column=2, value="").border = THIN_BORDER
                else:
                    val_cell = ws.cell(row=row, column=2, value=p.value)
                    if isinstance(p.value, float):
                        val_cell.number_format = "0.000000"
                    val_cell.border = THIN_BORDER
                ws.cell(row=row, column=3, value=p.unit).border = THIN_BORDER
                row += 1

            row += 1  # spacer

        # ── Area results section ──
        area_metrics = [m for m in (metrics or []) if self._is_area_metric(m)]
        if area_metrics:
            section_cell = ws.cell(row=row, column=1, value="Area Results")
            section_cell.font = Font(bold=True, size=13, color="2F5597")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.row_dimensions[row].height = 24
            row += 1

            for col_offset, header in enumerate(["Metric", "Value", "Unit"]):
                cell = ws.cell(row=row, column=1 + col_offset, value=header)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for m in area_metrics:
                name_cell = ws.cell(row=row, column=1, value=m.name)
                name_cell.font = Font(bold=True, size=11)
                name_cell.border = THIN_BORDER
                if isinstance(m.value, float | np.floating) and not np.isfinite(float(m.value)):
                    ws.cell(row=row, column=2, value="").border = THIN_BORDER
                else:
                    val_cell = ws.cell(row=row, column=2, value=m.value)
                    if isinstance(m.value, float):
                        val_cell.number_format = "0.000000"
                    val_cell.border = THIN_BORDER
                ws.cell(row=row, column=3, value=m.unit).border = THIN_BORDER
                row += 1

            row += 1  # spacer

        # ── Derived columns section ──
        section_cell = ws.cell(row=row, column=1, value="Derived Column Definitions")
        section_cell.font = Font(bold=True, size=13, color="2F5597")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=calc_col_count,
        )
        ws.row_dimensions[row].height = 24
        row += 1

        headers = ["Name", "Formula", "Unit", "Description", "Referenced Columns", "Status"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for dc in derived_columns:
            is_active = dc.enabled
            row_fill = active_fill if is_active else disabled_fill
            status = "Active" if is_active else "Disabled"

            for col_idx, value in [
                (1, dc.name), (2, dc.formula), (3, dc.unit),
                (4, dc.description), (5, dc.dependencies), (6, status),
            ]:
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.fill = row_fill
                if col_idx == 1:
                    cell.font = Font(bold=True, size=11)
            row += 1

        # ── Column widths – optimized for print ──
        col_widths = [22, 50, 12, 30, 30, 12]
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Print setup
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        logger.debug(f"Calculations sheet: {len(derived_columns)} derived columns")

    def _build_analysis_sheet(
        self,
        wb: Workbook,
        metrics: list[AnalysisMetricSpec],
        parameters: list[ParameterSpec],
        project_name: str = "Report",
    ) -> None:
        """Sheet 3: Analysis - print-friendly layout with title, grouped metrics, parameters."""
        ws = wb.create_sheet("Analysis")
        metric_col_count = 6
        area_highlight_fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )

        # ── Title row ──
        row = 1
        title_cell = ws.cell(row=row, column=1, value=f"Analysis Summary — {project_name}")
        title_cell.font = Font(bold=True, size=16, color="2F5597")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=metric_col_count,
        )
        ws.row_dimensions[row].height = 32
        row += 2  # blank spacer row

        # ── Parameters section (compact, before metrics so it's visible at top) ──
        if parameters:
            section_cell = ws.cell(row=row, column=1, value="Parameters")
            section_cell.font = Font(bold=True, size=13, color="2F5597")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.row_dimensions[row].height = 24
            row += 1

            for col_offset, header in enumerate(["Name", "Value", "Unit"]):
                cell = ws.cell(row=row, column=1 + col_offset, value=header)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for p in parameters:
                name_cell = ws.cell(row=row, column=1, value=p.name)
                name_cell.border = THIN_BORDER
                name_cell.font = Font(size=11)

                if isinstance(p.value, float | np.floating) and not np.isfinite(float(p.value)):
                    val_cell = ws.cell(row=row, column=2, value="")
                else:
                    val_cell = ws.cell(row=row, column=2, value=p.value)
                    if isinstance(p.value, float):
                        val_cell.number_format = "0.000000"
                val_cell.border = THIN_BORDER

                ws.cell(row=row, column=3, value=p.unit).border = THIN_BORDER
                row += 1

            row += 1  # spacer

        # ── Metrics section – grouped by chart ──
        section_cell = ws.cell(row=row, column=1, value="Analysis Metrics")
        section_cell.font = Font(bold=True, size=13, color="2F5597")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=metric_col_count,
        )
        ws.row_dimensions[row].height = 24
        row += 1

        metric_headers = ["Chart", "Metric", "Value", "Unit", "X Column", "Y Column"]
        for col_idx, header in enumerate(metric_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Group metrics by chart_title to insert visual separators
        current_chart: str | None = None
        for m in metrics:
            if m.chart_title != current_chart and current_chart is not None:
                # Light separator row between chart groups
                for c in range(1, metric_col_count + 1):
                    sep_cell = ws.cell(row=row, column=c, value="")
                    sep_cell.fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )
                ws.row_dimensions[row].height = 6
                row += 1
            current_chart = m.chart_title

            is_area = self._is_area_metric(m)
            metric_font = Font(bold=True, size=11) if is_area else Font(size=11)
            row_fill = area_highlight_fill if is_area else None

            chart_cell = ws.cell(row=row, column=1, value=m.chart_title)
            chart_cell.border = THIN_BORDER
            chart_cell.font = metric_font
            if row_fill:
                chart_cell.fill = row_fill

            name_cell = ws.cell(row=row, column=2, value=m.name)
            name_cell.border = THIN_BORDER
            name_cell.font = metric_font
            if row_fill:
                name_cell.fill = row_fill

            if isinstance(m.value, float | np.floating) and not np.isfinite(float(m.value)):
                val_cell = ws.cell(row=row, column=3, value="")
            else:
                val_cell = ws.cell(row=row, column=3, value=m.value)
                if isinstance(m.value, float):
                    val_cell.number_format = "0.000000"
            val_cell.border = THIN_BORDER
            val_cell.font = metric_font
            if row_fill:
                val_cell.fill = row_fill

            for col_idx, value in [(4, m.unit), (5, m.x_column), (6, m.y_column)]:
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if row_fill:
                    cell.fill = row_fill

            row += 1

        # ── Column widths – optimized for A4/Letter print ──
        col_widths = [22, 24, 18, 12, 18, 18]
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Print setup: fit columns to one page width
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        logger.debug(f"Analysis sheet: {len(metrics)} metrics, {len(parameters)} parameters")

    def _build_charts_sheet(
        self,
        wb: Workbook,
        chart_images: list[tuple],
        parameters: list[ParameterSpec] | None = None,
        metrics: list[AnalysisMetricSpec] | None = None,
    ) -> None:
        """Sheet 4: Charts - print-friendly layout matching Analysis/Calculations."""
        ws = wb.create_sheet("Charts")

        IMAGE_WIDTH = 800
        IMAGE_HEIGHT = 500
        ROWS_PER_CHART = 36
        row = 1

        # ── Title row ──
        title_cell = ws.cell(row=row, column=1, value="Charts")
        title_cell.font = Font(bold=True, size=16, color="2F5597")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 32
        row += 2  # blank spacer

        # ── Parameters section ──
        params = parameters or []
        if params:
            section_cell = ws.cell(row=row, column=1, value="Parameters")
            section_cell.font = Font(bold=True, size=13, color="2F5597")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.row_dimensions[row].height = 24
            row += 1

            for col_offset, header in enumerate(["Name", "Value", "Unit"]):
                cell = ws.cell(row=row, column=1 + col_offset, value=header)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for p in params:
                ws.cell(row=row, column=1, value=p.name).border = THIN_BORDER
                if isinstance(p.value, float | np.floating) and not np.isfinite(float(p.value)):
                    ws.cell(row=row, column=2, value="").border = THIN_BORDER
                else:
                    val_cell = ws.cell(row=row, column=2, value=p.value)
                    if isinstance(p.value, float):
                        val_cell.number_format = "0.000000"
                    val_cell.border = THIN_BORDER
                ws.cell(row=row, column=3, value=p.unit).border = THIN_BORDER
                row += 1

            row += 1  # spacer

        # ── Area Results section ──
        area_metrics = [m for m in (metrics or []) if self._is_area_metric(m)]
        if area_metrics:
            section_cell = ws.cell(row=row, column=1, value="Area Results")
            section_cell.font = Font(bold=True, size=13, color="2F5597")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.row_dimensions[row].height = 24
            row += 1

            for col_offset, header in enumerate(["Metric", "Value", "Unit"]):
                cell = ws.cell(row=row, column=1 + col_offset, value=header)
                cell.font = HEADER_FONT
                cell.fill = PARAM_HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1

            for m in area_metrics:
                name_cell = ws.cell(row=row, column=1, value=m.name)
                name_cell.font = Font(bold=True, size=11)
                name_cell.border = THIN_BORDER
                if isinstance(m.value, float | np.floating) and not np.isfinite(float(m.value)):
                    ws.cell(row=row, column=2, value="").border = THIN_BORDER
                else:
                    val_cell = ws.cell(row=row, column=2, value=m.value)
                    if isinstance(m.value, float):
                        val_cell.number_format = "0.000000"
                    val_cell.border = THIN_BORDER
                ws.cell(row=row, column=3, value=m.unit).border = THIN_BORDER
                row += 1

            row += 1  # spacer

        # ── Chart Figures section ──
        section_cell = ws.cell(row=row, column=1, value="Chart Figures")
        section_cell.font = Font(bold=True, size=13, color="2F5597")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 24
        row += 1

        for chart_title, chart_buf in chart_images:
            try:
                cell = ws.cell(row=row, column=1, value=chart_title)
                cell.font = Font(bold=True, size=12)
                ws.row_dimensions[row].height = 22
                row += 1

                chart_buf.seek(0)
                img = XlImage(chart_buf)
                img.width = IMAGE_WIDTH
                img.height = IMAGE_HEIGHT
                ws.add_image(img, f"A{row}")
                row += ROWS_PER_CHART

            except Exception as e:
                logger.warning(f"Chart '{chart_title}' failed: {e}")
                ws.cell(row=row, column=1, value=f"Chart could not be generated: {e}")
                row += 3

        ws.column_dimensions["A"].width = 100  # ~5 cell width for chart column
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        logger.debug(f"Charts sheet: {len(chart_images)} charts embedded")
