"""
Unit tests for src/services/xlsx_report_builder.py
"""

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook
from src.services.xlsx_report_builder import (
    AnalysisMetricSpec,
    DerivedColumnSpec,
    ParameterSpec,
    XlsxReportBuilder,
)


@pytest.fixture
def sample_df():
    """Sample DataFrame for testing."""
    return pd.DataFrame(
        {
            "Time": [0.0, 0.1, 0.2, 0.3, 0.4],
            "Load": [0.0, 10.0, 20.0, 15.0, 5.0],
            "Extension": [0.0, 0.01, 0.02, 0.03, 0.04],
            "Stress": [0.0, 100.0, 200.0, 150.0, 50.0],
        }
    )


@pytest.fixture
def column_units():
    return {"Load": "N", "Extension": "mm", "Stress": "MPa"}


@pytest.fixture
def derived_columns():
    return [
        DerivedColumnSpec(
            name="Stress",
            formula="[Load] / ([width] * [thickness])",
            unit="MPa",
            description="Engineering stress",
            dependencies="Load, width, thickness",
            enabled=True,
        ),
    ]


@pytest.fixture
def metrics():
    return [
        AnalysisMetricSpec(
            chart_id="chart-001",
            chart_title="Stress-Strain",
            name="Max Stress",
            value=200.0,
            unit="MPa",
            x_column="Extension",
            y_column="Stress",
        ),
        AnalysisMetricSpec(
            chart_id="chart-001",
            chart_title="Stress-Strain",
            name="Total Area",
            value=5.5,
            unit="MJ/m^3",
            x_column="Extension",
            y_column="Stress",
        ),
    ]


@pytest.fixture
def parameters():
    return [
        ParameterSpec(name="length", value=50.0, unit="mm"),
        ParameterSpec(name="width", value=12.5, unit="mm"),
        ParameterSpec(name="thickness", value=2.0, unit="mm"),
    ]


def _build_report(sample_df, column_units, derived_columns, metrics, parameters, **kwargs):
    builder = XlsxReportBuilder()
    return builder.build(
        df=sample_df,
        column_units=column_units,
        derived_columns=derived_columns,
        metrics=metrics,
        parameters=parameters,
        chart_images=kwargs.get("chart_images", []),
        project_name=kwargs.get("project_name", "Test"),
    )


class TestXlsxReportBuilder:
    """Tests for XlsxReportBuilder."""

    def test_builds_workbook_with_four_sheets(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Should create workbook with 4 sheets."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 4
        assert wb.sheetnames == ["Data", "Calculations", "Analysis", "Charts"]

    def test_data_sheet_has_correct_headers(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Data sheet row 2 should have column headers (row 1 is title)."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        wb = load_workbook(output)
        ws = wb["Data"]
        headers = [ws.cell(row=2, column=i).value for i in range(1, 5)]
        assert headers == ["Time", "Load", "Extension", "Stress"]

    def test_data_sheet_has_units_row(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Data sheet row 3 should have units."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        wb = load_workbook(output)
        ws = wb["Data"]
        units = [ws.cell(row=3, column=i).value for i in range(1, 5)]
        assert units[1] == "N"
        assert units[2] == "mm"
        assert units[3] == "MPa"

    def test_data_sheet_has_correct_row_count(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Data should start at row 4 with 5 data rows."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        wb = load_workbook(output)
        ws = wb["Data"]
        assert ws.cell(row=4, column=1).value == 0.0
        assert ws.cell(row=8, column=2).value == 5.0

    def test_calculations_sheet_has_derived_columns(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Calculations sheet should list derived column definitions."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        wb = load_workbook(output)
        ws = wb["Calculations"]
        # Find "Stress" derived column by searching (layout is sectioned)
        stress_found = False
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Stress":
                assert ws.cell(row=r, column=2).value == "[Load] / ([width] * [thickness])"
                stress_found = True
                break
        assert stress_found, "Stress derived column not found in Calculations sheet"

    def test_empty_data(self):
        """Should handle empty DataFrame."""
        output = _build_report(pd.DataFrame(), {}, [], [], [])
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 4

    def test_returns_bytesio(self, sample_df, column_units, derived_columns, metrics, parameters):
        """Should return a BytesIO object."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        assert isinstance(output, BytesIO)
        assert len(output.getvalue()) > 0


class TestAnalysisSheetLayout:
    """Tests for the redesigned print-friendly Analysis sheet."""

    def _get_analysis_ws(self, sample_df, column_units, derived_columns, metrics, parameters):
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        return load_workbook(output)["Analysis"]

    def test_title_row(self, sample_df, column_units, derived_columns, metrics, parameters):
        """Row 1 should contain the report title."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        assert "Analysis Summary" in ws.cell(row=1, column=1).value

    def test_parameters_section_before_metrics(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Parameters section should appear before metrics with correct header."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        # Row 3 = "Parameters" section header (row 2 is spacer)
        assert ws.cell(row=3, column=1).value == "Parameters"

    def test_parameter_headers(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Parameter table should have Name/Value/Unit headers."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        # Row 4 = parameter headers
        assert ws.cell(row=4, column=1).value == "Name"
        assert ws.cell(row=4, column=2).value == "Value"
        assert ws.cell(row=4, column=3).value == "Unit"

    def test_parameter_values(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Parameter data should be written after headers."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        # Row 5, 6, 7 = parameter data (length, width, thickness)
        assert ws.cell(row=5, column=1).value == "length"
        assert ws.cell(row=5, column=2).value == 50.0
        assert ws.cell(row=5, column=3).value == "mm"
        assert ws.cell(row=6, column=1).value == "width"
        assert ws.cell(row=7, column=1).value == "thickness"

    def test_metrics_section_header(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Metrics section should have a clear section header."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        # 3 params: rows 3(header), 4(col headers), 5-7(data), 8(spacer), 9 = Metrics header
        assert ws.cell(row=9, column=1).value == "Analysis Metrics"

    def test_metric_column_headers(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Metrics table should have proper column headers."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        headers = [ws.cell(row=10, column=i).value for i in range(1, 7)]
        assert headers == ["Chart", "Metric", "Value", "Unit", "X Column", "Y Column"]

    def test_metric_data(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Metrics data should appear after the column headers."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        assert ws.cell(row=11, column=1).value == "Stress-Strain"
        assert ws.cell(row=11, column=2).value == "Max Stress"
        assert ws.cell(row=11, column=3).value == 200.0

    def test_area_metric_highlighted(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Area-related metrics should have a highlight fill."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        # Row 12 is "Total Area" (area metric) — should be highlighted
        assert ws.cell(row=12, column=2).value == "Total Area"
        fill = ws.cell(row=12, column=2).fill
        assert fill.start_color.rgb is not None

    def test_no_parameters_skips_section(self, sample_df, column_units, derived_columns, metrics):
        """When no parameters, metrics section should start earlier."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, [])
        ws = load_workbook(output)["Analysis"]
        # With no parameters: row 1 = title, row 2 = spacer, row 3 = "Analysis Metrics"
        assert ws.cell(row=3, column=1).value == "Analysis Metrics"

    def test_column_widths_set(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Column widths should be set for print readability."""
        ws = self._get_analysis_ws(sample_df, column_units, derived_columns, metrics, parameters)
        assert ws.column_dimensions["A"].width >= 20
        assert ws.column_dimensions["C"].width >= 16


class TestCalculationsSheetLayout:
    """Tests for the redesigned Calculations sheet (matching Analysis tab style)."""

    def _get_calc_ws(self, sample_df, column_units, derived_columns, metrics, parameters):
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        return load_workbook(output)["Calculations"]

    def test_title_row(self, sample_df, column_units, derived_columns, metrics, parameters):
        """Row 1 should contain the sheet title."""
        ws = self._get_calc_ws(sample_df, column_units, derived_columns, metrics, parameters)
        assert "Derived Columns" in ws.cell(row=1, column=1).value

    def test_parameters_section(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Parameters section should appear with correct data."""
        ws = self._get_calc_ws(sample_df, column_units, derived_columns, metrics, parameters)
        # Row 3 = "Parameters" section header
        assert ws.cell(row=3, column=1).value == "Parameters"
        # Row 4 = headers, Row 5+ = data
        assert ws.cell(row=4, column=1).value == "Name"
        assert ws.cell(row=5, column=1).value == "length"
        assert ws.cell(row=5, column=2).value == 50.0

    def test_area_results_section(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Area results section should appear between params and derived columns."""
        ws = self._get_calc_ws(sample_df, column_units, derived_columns, metrics, parameters)
        area_found = False
        for r in range(1, 20):
            if ws.cell(row=r, column=1).value == "Area Results":
                area_found = True
                break
        assert area_found

    def test_derived_columns_section_header(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Derived columns section should have a clear section header."""
        ws = self._get_calc_ws(sample_df, column_units, derived_columns, metrics, parameters)
        section_found = False
        for r in range(1, 25):
            if ws.cell(row=r, column=1).value == "Derived Column Definitions":
                section_found = True
                break
        assert section_found

    def test_active_column_has_status_fill(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Active derived columns should have a fill color."""
        ws = self._get_calc_ws(sample_df, column_units, derived_columns, metrics, parameters)
        for r in range(1, 25):
            if ws.cell(row=r, column=1).value == "Stress":
                fill = ws.cell(row=r, column=1).fill
                assert fill.start_color.rgb is not None
                assert ws.cell(row=r, column=6).value == "Active"
                break

    def test_no_params_skips_section(self, sample_df, column_units, derived_columns, metrics):
        """When no parameters, derived columns section starts earlier."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, [])
        ws = load_workbook(output)["Calculations"]
        # Row 1 = title, row 2 = spacer, then area or derived columns
        found_params = False
        for r in range(1, 15):
            if ws.cell(row=r, column=1).value == "Parameters":
                found_params = True
        assert not found_params


class TestChartsSheetLayout:
    """Tests for the redesigned Charts sheet (matching Analysis/Calculations style)."""

    def test_charts_sheet_has_title(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Charts sheet should have a title row."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        ws = load_workbook(output)["Charts"]
        assert ws.cell(row=1, column=1).value == "Charts"

    def test_charts_sheet_has_parameters_section(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Charts sheet should show Parameters section."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        ws = load_workbook(output)["Charts"]
        param_found = False
        for r in range(1, 20):
            if ws.cell(row=r, column=1).value == "Parameters":
                param_found = True
                break
        assert param_found

    def test_charts_sheet_has_area_results_section(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Charts sheet should show Area Results section."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        ws = load_workbook(output)["Charts"]
        area_found = False
        for r in range(1, 25):
            if ws.cell(row=r, column=1).value == "Area Results":
                area_found = True
                break
        assert area_found

    def test_charts_sheet_has_chart_figures_section(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Charts sheet should have Chart Figures section header."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        ws = load_workbook(output)["Charts"]
        section_found = False
        for r in range(1, 30):
            if ws.cell(row=r, column=1).value == "Chart Figures":
                section_found = True
                break
        assert section_found

    def test_charts_sheet_no_params_or_area_when_empty(self, sample_df, column_units, derived_columns):
        """Charts sheet has title and Chart Figures but no Parameters/Area when none provided."""
        non_area = [
            AnalysisMetricSpec(
                chart_id="c1", chart_title="T", name="Peak",
                value=10.0, unit="N", x_column="X", y_column="Y",
            ),
        ]
        output = _build_report(sample_df, column_units, derived_columns, non_area, [])
        ws = load_workbook(output)["Charts"]
        assert ws.cell(row=1, column=1).value == "Charts"
        param_found = any(
            ws.cell(row=r, column=1).value == "Parameters" for r in range(1, 15)
        )
        area_found = any(
            ws.cell(row=r, column=1).value == "Area Results" for r in range(1, 15)
        )
        assert not param_found
        assert not area_found


class TestDataSheetParameters:
    """Tests for parameters on the Data sheet."""

    def test_data_sheet_has_parameter_metadata(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Data sheet should have parameter metadata in a sidebar."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        ws = load_workbook(output)["Data"]
        # 4 data columns + 2-col gap = param starts at col 7
        assert ws.cell(row=2, column=7).value == "length"
        assert ws.cell(row=2, column=8).value == "width"

    def test_data_sheet_has_parameter_values(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Data sheet parameter values should be in row 4."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        ws = load_workbook(output)["Data"]
        assert ws.cell(row=4, column=7).value == 50.0
        assert ws.cell(row=4, column=8).value == 12.5

    def test_data_sheet_has_area_results_in_sidebar(
        self, sample_df, column_units, derived_columns, metrics, parameters
    ):
        """Data sheet should have area results in the sidebar (right side, near top)."""
        output = _build_report(sample_df, column_units, derived_columns, metrics, parameters)
        ws = load_workbook(output)["Data"]
        # Sidebar starts at column 7 (4 data cols + 2 gap). Area at row 6 (below params).
        assert ws.cell(row=6, column=7).value == "Area Results"
        assert ws.cell(row=7, column=7).value == "Metric"
        assert ws.cell(row=8, column=7).value == "Total Area"
        assert ws.cell(row=8, column=8).value == 5.5

    def test_data_sheet_no_area_section_when_no_area_metrics(
        self, sample_df, column_units, derived_columns, parameters
    ):
        """Data sheet should not have area section when no area metrics exist."""
        non_area = [
            AnalysisMetricSpec(
                chart_id="c1", chart_title="T", name="Peak",
                value=10.0, unit="N", x_column="X", y_column="Y",
            ),
        ]
        output = _build_report(sample_df, column_units, derived_columns, non_area, parameters)
        ws = load_workbook(output)["Data"]
        # Check sidebar area for no "Area Results"
        for r in range(4, 15):
            assert ws.cell(row=r, column=7).value != "Area Results"
