import math
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from src.api.server import app

client = TestClient(app)


def _upload_dataset_id(rows: list[dict[str, float | int]]) -> str:
    if not rows:
        raise ValueError("rows must not be empty")
    headers = list(rows[0].keys())
    lines = [",".join(headers)]
    for row in rows:
        values = [str(row.get(header, "")) for header in headers]
        lines.append(",".join(values))
    csv_content = "\n".join(lines) + "\n"
    response = client.post(
        "/api/v3/datasets/upload",
        files={"file": ("metrics.csv", csv_content, "text/csv")},
    )
    assert response.status_code == 200, response.text
    dataset_id = response.json().get("dataset_id")
    assert isinstance(dataset_id, str) and dataset_id
    return dataset_id


def test_request_id_header_roundtrip():
    req_id = "test-request-id-123"
    response = client.get("/api/v3/health", headers={"X-Request-ID": req_id})
    assert response.status_code == 200
    assert response.headers.get("x-request-id") == req_id


def test_request_id_header_generated_when_missing():
    response = client.get("/api/v3/health")
    assert response.status_code == 200
    assert response.headers.get("x-request-id")


def test_request_id_in_request_validation_error():
    req_id = "request-validation-id"
    response = client.post("/api/v3/datasets/upload", headers={"X-Request-ID": req_id})
    assert response.status_code == 422, response.text
    payload = response.json()
    assert payload.get("code") == "REQUEST_VALIDATION_ERROR"
    assert payload.get("request_id") == req_id


def test_request_id_in_http_exception_error():
    req_id = "http-exception-id"
    dataset_id = _upload_dataset_id([{"X": 1}, {"X": 2}, {"X": 3}])
    payload = {
        "dataset_id": dataset_id,
        "parameters": {},
        "units": {"X": ""},
        "column_mapping": {},
        "project_name": "HTTPException Test",
        "operations": [],
        "initial_results": {},
        "derived_parameters": [],
        "user_formulas": [
            {
                "name": "Broken",
                "formula": "[Missing] + 1",
                "unit": "",
                "description": "",
                "enabled": True,
            }
        ],
    }

    response = client.post("/api/v3/processing/run", json=payload, headers={"X-Request-ID": req_id})
    assert response.status_code == 422, response.text
    body = response.json()
    assert body.get("code") == "HTTP_ERROR"
    assert body.get("request_id") == req_id


def test_calculate_metrics_area_total_equals_sum_of_selected_regions():
    payload = {
        "data": [
            {"X": -2, "Y": 3},
            {"X": -1, "Y": 2},
            {"X": 0, "Y": 1},
            {"X": 1, "Y": -1},
            {"X": 2, "Y": -2},
        ],
        "charts": [
            {
                "id": "chart-1",
                "title": "Area Chart",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "area",
                "area_spec": {
                    "mode": "total",
                    "baseline": 0.0,
                    "x_column": "X",
                    "y_column": "Y",
                    "label": "Area",
                },
                "baseline_spec": {
                    "x_baseline": 0.0,
                    "y_baseline": 0.0,
                    "regions": ["top-left", "bottom-right"],
                },
            }
        ],
    }

    response = client.post("/api/v3/charts/metrics", json=payload)
    assert response.status_code == 200, response.text

    metrics = response.json()["metrics"]
    assert len(metrics) == 1
    metric = metrics[0]

    area_by_region = metric["area_by_region"]
    assert "top-left" in area_by_region
    assert "bottom-right" in area_by_region

    expected_total = area_by_region["top-left"] + area_by_region["bottom-right"]
    assert metric["area_total"] == pytest.approx(expected_total)


def test_calculate_metrics_non_area_charts_return_no_metrics():
    payload = {
        "data": [
            {"X": 0, "Y1": 1, "Y2": -10},
            {"X": 1, "Y1": 2, "Y2": -10},
            {"X": 2, "Y1": 3, "Y2": 5},
            {"X": 3, "Y1": 4, "Y2": 6},
        ],
        "charts": [
            {
                "id": "chart-scope",
                "title": "Scoped Chart",
                "x_column": "X",
                "y_columns": ["Y1", "Y2"],
                "chart_type": "line",
                "scope": {
                    "mode": "range",
                    "y_min": 0,
                },
            }
        ],
    }

    response = client.post("/api/v3/charts/metrics", json=payload)
    assert response.status_code == 200, response.text
    assert response.json()["metrics"] == []


def test_calculate_metrics_accepts_run_id_from_processing_session():
    dataset_id = _upload_dataset_id(
        [
            {"X": 0, "Y": 1},
            {"X": 1, "Y": 3},
            {"X": 2, "Y": 2},
        ]
    )
    process_payload = {
        "project_name": "Metrics Run ID",
        "dataset_id": dataset_id,
    }
    process_response = client.post("/api/v3/processing/run", json=process_payload)
    assert process_response.status_code == 200, process_response.text
    run_id = process_response.json()["run_id"]

    metrics_payload = {
        "run_id": run_id,
        "charts": [
            {
                "id": "chart-run-id",
                "title": "Run ID Metrics",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "area",
                "area_spec": {
                    "mode": "total",
                    "baseline": 0.0,
                    "x_column": "X",
                    "y_column": "Y",
                    "label": "RunIdArea",
                },
            }
        ],
    }
    response = client.post("/api/v3/charts/metrics", json=metrics_payload)
    assert response.status_code == 200, response.text
    metrics = response.json()["metrics"]
    assert len(metrics) == 1
    assert metrics[0]["area_total"] == pytest.approx(4.5)


def test_calculate_metrics_skips_unlabeled_area_metrics():
    payload = {
        "data": [
            {"X": 0, "Y": 1},
            {"X": 1, "Y": 3},
        ],
        "charts": [
            {
                "id": "chart-unlabeled-area",
                "title": "Unlabeled Area",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "area",
                "area_spec": {
                    "mode": "total",
                    "baseline": 0.0,
                    "x_column": "X",
                    "y_column": "Y",
                    "label": " ",
                },
            }
        ],
    }

    response = client.post("/api/v3/charts/metrics", json=payload)
    assert response.status_code == 200, response.text
    assert response.json()["metrics"] == []


def test_calculate_metrics_handles_non_finite_baseline_without_500():
    payload = {
        "data": [
            {"X": 0, "Y": 1},
            {"X": 1, "Y": 3},
            {"X": 2, "Y": 2},
        ],
        "charts": [
            {
                "id": "chart-nan-baseline",
                "title": "Non-finite baseline",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "area",
                "area_spec": {
                    "mode": "total",
                    "baseline": "NaN",
                    "x_column": "X",
                    "y_column": "Y",
                    "label": "Area",
                },
            }
        ],
    }

    response = client.post("/api/v3/charts/metrics", json=payload)
    assert response.status_code == 200, response.text
    body = response.json()
    assert "metrics" in body
    for metric in body["metrics"]:
        area_total = metric.get("area_total")
        if area_total is not None:
            assert math.isfinite(area_total)


def test_export_v2_accepts_dataset_id_without_raw_data():
    rows = [
        {"X": 0, "Y": 1},
        {"X": 1, "Y": 3},
        {"X": 2, "Y": 2},
    ]
    dataset_id = _upload_dataset_id(rows)

    export_payload = {
        "project_name": "Export via Run ID",
        "dataset_id": dataset_id,
        "selected_columns": ["X", "Y"],
        "parameters": {},
        "units": {"X": "s", "Y": "N"},
        "charts": [],
        "metrics": [],
    }
    response = client.post("/api/v3/reports/xlsx", json=export_payload)
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_export_v2_chart_generation_does_not_depend_on_selected_columns():
    rows = [
        {"X": 0, "Y": 1, "Other": 10},
        {"X": 1, "Y": 3, "Other": 20},
        {"X": 2, "Y": 2, "Other": 30},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "project_name": "Export Chart Test",
        "parameters": {},
        "units": {"X": "s", "Y": "N", "Other": ""},
        "selected_columns": ["Other"],
        "derived_columns": [],
        "metrics": [],
        "charts": [
            {
                "id": "chart-export",
                "title": "Line Export",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "line",
            }
        ],
    }

    response = client.post("/api/v3/reports/xlsx", json=payload)
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    wb = load_workbook(BytesIO(response.content))
    charts_ws = wb["Charts"]
    assert len(charts_ws._images) == 1


def test_export_v2_renders_all_requested_charts_line_and_area():
    rows = [
        {"X": 0, "Y": 1},
        {"X": 1, "Y": 3},
        {"X": 2, "Y": 2},
        {"X": 3, "Y": 4},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "project_name": "Export Two Charts",
        "parameters": {},
        "units": {"X": "s", "Y": "N"},
        "selected_columns": ["X", "Y"],
        "derived_columns": [],
        "metrics": [],
        "charts": [
            {
                "id": "chart-line",
                "title": "Line Chart",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "line",
                "baseline_spec": {
                    "x_baseline": 0.0,
                    "y_baseline": 0.0,
                    "regions": ["top-right"],
                },
            },
            {
                "id": "chart-area",
                "title": "Area Chart",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "area",
                "area_spec": {
                    "mode": "positive",
                    "baseline": 0.0,
                    "x_column": "X",
                    "y_column": "Y",
                    "label": "Positive Area",
                },
                "baseline_spec": {
                    "x_baseline": 0.0,
                    "y_baseline": 0.0,
                    "regions": ["top-right"],
                },
            },
        ],
    }

    response = client.post("/api/v3/reports/xlsx", json=payload)
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))
    charts_ws = wb["Charts"]
    assert len(charts_ws._images) == 2


def test_export_v2_renders_line_scatter_area_with_customizations():
    rows = [
        {"X": -2, "Y1": -1, "Y2": 4},
        {"X": -1, "Y1": 1, "Y2": 3},
        {"X": 0, "Y1": 2, "Y2": 2},
        {"X": 1, "Y1": 3, "Y2": 1},
        {"X": 2, "Y1": 4, "Y2": 0},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "project_name": "Export Three Chart Types",
        "parameters": {},
        "units": {"X": "s", "Y1": "N", "Y2": "N"},
        "selected_columns": ["X", "Y1", "Y2"],
        "derived_columns": [],
        "metrics": [],
        "charts": [
            {
                "id": "chart-line-custom",
                "title": "Line Custom",
                "x_column": "X",
                "y_columns": ["Y1"],
                "chart_type": "line",
                "line_color": "#FF0000",
                "line_width": 3.0,
                "baseline_spec": {
                    "x_baseline": 0.0,
                    "y_baseline": 0.0,
                    "regions": ["top-right"],
                },
            },
            {
                "id": "chart-scatter-custom",
                "title": "Scatter Custom",
                "x_column": "X",
                "y_columns": ["Y2"],
                "chart_type": "scatter",
                "line_color": "#00AA88",
                "marker_size": 28.0,
            },
            {
                "id": "chart-area-custom",
                "title": "Area Custom",
                "x_column": "X",
                "y_columns": ["Y1"],
                "chart_type": "area",
                "line_color": "#2244EE",
                "fill_color": "#88CCFF",
                "fill_opacity": 0.5,
                "area_spec": {
                    "mode": "total",
                    "baseline": 0.0,
                    "x_column": "X",
                    "y_column": "Y1",
                    "label": "Custom Area",
                },
                "baseline_spec": {
                    "x_baseline": 0.0,
                    "y_baseline": 0.0,
                    "regions": [],
                },
            },
        ],
    }

    response = client.post("/api/v3/reports/xlsx", json=payload)
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))
    charts_ws = wb["Charts"]
    assert len(charts_ws._images) == 3


def test_export_v2_includes_derived_columns_with_order_and_group_color():
    rows = [
        {"X": 1, "Y": 2},
        {"X": 2, "Y": 3},
        {"X": 3, "Y": 5},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "project_name": "Export Derived Test",
        "parameters": {},
        "units": {"X": "s", "Y": "N"},
        "selected_columns": ["X", "Y"],
        # Intentionally omit user_formulas: backend should still evaluate via derived_columns.
        "derived_columns": [
            {
                "id": "dc-1",
                "name": "XY_sum",
                "formula": "[X] + [Y]",
                "unit": "N",
                "description": "sum",
                "dependencies": ["X", "Y"],
                "enabled": True,
            }
        ],
        "column_layout": {
            "column_order": ["X", "XY_sum", "Y"],
            "separator_indices": [1],
            "separator_color": "#FF00AA",
            "linked_groups": [],
            "matching_groups": [],
        },
        "column_colors": {
            "XY_sum": "AABBCC",
        },
        "metrics": [],
        "charts": [
            {
                "id": "chart-derived",
                "title": "Derived Line",
                "x_column": "X",
                "y_columns": ["XY_sum"],
                "chart_type": "line",
            }
        ],
    }

    response = client.post("/api/v3/reports/xlsx", json=payload)
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))

    data_ws = wb["Data"]

    # Expect order: X, XY_sum, <separator>, Y
    assert data_ws.cell(row=1, column=1).value == "X"
    assert data_ws.cell(row=1, column=2).value == "XY_sum"
    assert data_ws.cell(row=1, column=3).value in ("", None)
    assert data_ws.cell(row=1, column=4).value == "Y"

    # Derived column values should be present and computed.
    assert data_ws.cell(row=3, column=2).value == pytest.approx(3.0)
    assert data_ws.cell(row=4, column=2).value == pytest.approx(5.0)
    assert data_ws.cell(row=5, column=2).value == pytest.approx(8.0)

    # Group color applies to units/data (not header); openpyxl stores ARGB.
    unit_fill = data_ws.cell(row=2, column=2).fill.start_color.index or ""
    assert str(unit_fill).upper().endswith("AABBCC")

    # User-selected separator color should be applied to separator column.
    sep_fill = data_ws.cell(row=1, column=3).fill.start_color.index or ""
    assert str(sep_fill).upper().endswith("FF00AA")

    charts_ws = wb["Charts"]
    assert len(charts_ws._images) == 1


def test_export_v2_applies_matching_group_colors_without_column_colors_map():
    rows = [
        {"A": 1, "B": 2},
        {"A": 2, "B": 3},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "project_name": "Export Matching Group Colors",
        "parameters": {},
        "units": {"A": "mm", "B": "N"},
        "selected_columns": ["A", "B"],
        "derived_columns": [],
        "metrics": [],
        "charts": [],
        "column_layout": {
            "column_order": ["A", "B"],
            "separator_indices": [],
            "linked_groups": [],
            "matching_groups": [
                {
                    "id": "mg-a",
                    "name": "Group A",
                    "color": "#55AAEE",
                    "columns": ["A"],
                }
            ],
        },
        "column_colors": {},
    }

    response = client.post("/api/v3/reports/xlsx", json=payload)
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))
    data_ws = wb["Data"]

    # Header stays default blue; units/data should be group colored.
    unit_fill = data_ws.cell(row=2, column=1).fill.start_color.index or ""
    data_fill = data_ws.cell(row=3, column=1).fill.start_color.index or ""
    assert str(unit_fill).upper().endswith("55AAEE")
    assert str(data_fill).upper().endswith("55AAEE")


def test_export_v2_resolves_formula_display_label_aliases():
    rows = [
        {"Tensile stress": 2.0},
        {"Tensile stress": 3.5},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "project_name": "Export Alias Formula Test",
        "parameters": {},
        "units": {"Tensile stress": "MPa"},
        "selected_columns": ["Tensile stress"],
        "user_formulas": [
            {
                "name": "DoubleStress",
                "formula": "[True Stress] * 2",
                "unit": "MPa",
                "description": "Alias based",
                "enabled": True,
            }
        ],
        "derived_columns": [
            {
                "id": "dc-alias",
                "name": "DoubleStress",
                "formula": "[True Stress] * 2",
                "unit": "MPa",
                "description": "Alias based",
                "dependencies": ["Tensile stress"],
                "enabled": True,
            }
        ],
        "column_layout": {
            "column_order": ["Tensile stress", "DoubleStress"],
            "separator_indices": [],
            "linked_groups": [],
            "matching_groups": [],
        },
        "header_mapping": {
            "Tensile stress": "True Stress",
        },
        "metrics": [],
        "charts": [],
    }

    response = client.post("/api/v3/reports/xlsx", json=payload)
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))
    data_ws = wb["Data"]

    assert data_ws.cell(row=1, column=1).value == "True Stress"
    assert data_ws.cell(row=1, column=2).value == "DoubleStress"
    assert data_ws.cell(row=3, column=2).value == pytest.approx(4.0)
    assert data_ws.cell(row=4, column=2).value == pytest.approx(7.0)


def test_export_v2_keeps_request_metrics_when_charts_are_present():
    rows = [
        {"X": 0, "Y": 1},
        {"X": 1, "Y": 3},
        {"X": 2, "Y": 2},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "project_name": "Export Metrics Merge",
        "parameters": {},
        "units": {"X": "s", "Y": "N"},
        "selected_columns": ["X", "Y"],
        "derived_columns": [],
        "metrics": [
            {
                "id": "param-1",
                "chart_id": "",
                "chart_title": "",
                "name": "A0",
                "value": 12.5,
                "unit": "mm2",
                "x_column": "",
                "y_column": "",
            }
        ],
        "charts": [
            {
                "id": "chart-line",
                "title": "Line",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "line",
            }
        ],
    }

    response = client.post("/api/v3/reports/xlsx", json=payload)
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))
    analysis_ws = wb["Analysis"]

    assert analysis_ws.cell(row=2, column=2).value == "A0"
    assert analysis_ws.cell(row=2, column=3).value == pytest.approx(12.5)


def test_export_v2_includes_derived_parameter_metrics_without_frontend_metrics():
    rows = [
        {"X": 10, "Y": 1},
        {"X": 10, "Y": 3},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "project_name": "Export Derived Parameter Fallback",
        "parameters": {},
        "parameter_units": {"A0": "mm2"},
        "units": {"X": "s", "Y": "N"},
        "selected_columns": ["X", "Y"],
        "derived_columns": [],
        "derived_parameters": [
            {
                "name": "A0",
                "formula": "REF([X])",
            }
        ],
        "reference_index": 0,
        "metrics": [],
        "charts": [
            {
                "id": "chart-line",
                "title": "Line",
                "x_column": "X",
                "y_columns": ["Y"],
                "chart_type": "line",
            }
        ],
    }

    response = client.post("/api/v3/reports/xlsx", json=payload)
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))
    analysis_ws = wb["Analysis"]

    assert analysis_ws.cell(row=2, column=2).value == "A0"
    assert analysis_ws.cell(row=2, column=3).value == pytest.approx(10.0)
    assert analysis_ws.cell(row=7, column=1).value == "A0"
    assert analysis_ws.cell(row=7, column=2).value == pytest.approx(10.0)


def test_process_resolves_derived_parameter_alias_without_header_mapping():
    """Processing pipeline uses exact column names (strict mode — no fuzzy matching)."""
    rows = [
        {"Tensile stress": 2.0},
        {"Tensile stress": 3.5},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "parameters": {},
        "units": {"Tensile stress": "MPa"},
        "column_mapping": {},
        "project_name": "Alias Process Test",
        "operations": [],
        "initial_results": {},
        "derived_parameters": [{"name": "Tensile Strength", "formula": "[Tensile stress]"}],
        "user_formulas": [
            {
                "name": "DoubleStrength",
                "formula": "[Tensile Strength] * 2",
                "unit": "MPa",
                "description": "",
                "enabled": True,
            }
        ],
    }

    response = client.post("/api/v3/processing/run", json=payload)
    assert response.status_code == 200, response.text

    run_id = response.json()["run_id"]
    run_data = client.get(f"/api/v3/processing/runs/{run_id}/data")
    assert run_data.status_code == 200, run_data.text
    rows = run_data.json()["processed_data"]
    assert rows[0]["DoubleStrength"] == pytest.approx(4.0)
    assert rows[1]["DoubleStrength"] == pytest.approx(7.0)


def test_process_rejects_fuzzy_reference_in_strict_mode():
    """Fuzzy references like [True Stress] should NOT match [Tensile stress] in strict mode."""
    rows = [
        {"Tensile stress": 2.0},
        {"Tensile stress": 3.5},
    ]
    dataset_id = _upload_dataset_id(rows)
    payload = {
        "dataset_id": dataset_id,
        "parameters": {},
        "units": {"Tensile stress": "MPa"},
        "column_mapping": {},
        "project_name": "Strict Mode Test",
        "operations": [],
        "initial_results": {},
        "derived_parameters": [{"name": "Tensile Strength", "formula": "[True Stress]"}],
        "user_formulas": [],
    }

    response = client.post("/api/v3/processing/run", json=payload)
    assert response.status_code == 422, "Fuzzy references should fail in strict mode"
